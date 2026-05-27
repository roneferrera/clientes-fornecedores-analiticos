# app_cnpj.py
import io
import re
import time
import requests
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ==============================
# VERSÃO
# ==============================
VERSAO = "V1.0"

# ==============================
# TEMA TR
# ==============================
def apply_tr_theme():
    st.markdown("""
        <style>
        html, body, [class*="css"] {
            font-family: 'Segoe UI', 'Arial', sans-serif;
            color: #444444;
        }
        h1, h2, h3 {
            color: #FF8000;
            font-weight: 700;
        }
        section[data-testid="stSidebar"] {
            background-color: #444444;
            color: #FFFFFF;
        }
        section[data-testid="stSidebar"] * {
            color: #FFFFFF !important;
        }
        .stButton > button {
            background-color: #FF8000;
            color: #FFFFFF;
            border: none;
            border-radius: 4px;
            font-weight: bold;
        }
        .stButton > button:hover {
            background-color: #D64001;
            color: #FFFFFF;
        }
        .stDownloadButton > button {
            background-color: #FF8000;
            color: #FFFFFF;
            border: none;
            border-radius: 4px;
            font-weight: bold;
        }
        .stDownloadButton > button:hover {
            background-color: #D64001;
            color: #FFFFFF;
        }
        hr { border-color: #FF8000; }
        [data-testid="metric-container"] {
            background-color: #E9E9E9;
            border-left: 4px solid #FF8000;
            border-radius: 4px;
            padding: 10px;
        }
        .instrucoes-box {
            background-color: #E9E9E9;
            border-left: 4px solid #FF8000;
            border-radius: 4px;
            padding: 16px 20px;
            margin: 12px 0;
            color: #444444;
            font-family: 'Segoe UI', Arial, sans-serif;
        }
        .instrucoes-box h4 {
            color: #FF8000;
            margin-top: 14px;
            margin-bottom: 6px;
        }
        .instrucoes-box h4:first-child { margin-top: 0; }
        </style>
    """, unsafe_allow_html=True)


# ==============================
# UTILITÁRIOS
# ==============================
def texto(v):
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none") else s


def so_numeros(v):
    return re.sub(r"\D", "", str(v or ""))


def formatar_cnpj(cnpj: str) -> str:
    n = so_numeros(cnpj).zfill(14)
    return f"{n[:2]}.{n[2:5]}.{n[5:8]}/{n[8:12]}-{n[12:14]}"


def formatar_cep(cep: str) -> str:
    n = so_numeros(cep).zfill(8)
    return f"{n[:5]}-{n[5:]}"


def validar_cnpj(cnpj: str) -> bool:
    n = so_numeros(cnpj)
    if len(n) != 14 or len(set(n)) == 1:
        return False
    def calc(digits, pesos):
        s = sum(int(d) * p for d, p in zip(digits, pesos))
        r = s % 11
        return 0 if r < 2 else 11 - r
    return (
        calc(n[:12], [5,4,3,2,9,8,7,6,5,4,3,2]) == int(n[12]) and
        calc(n[:13], [6,5,4,3,2,9,8,7,6,5,4,3,2]) == int(n[13])
    )


def extrair_telefone(raw: str) -> str:
    if not raw:
        return ""
    nums = so_numeros(str(raw))
    if len(nums) >= 10:
        ddd    = nums[:2]
        numero = nums[2:]
        if len(numero) == 9:
            return f"({ddd}) {numero[:5]}-{numero[5:]}"
        elif len(numero) == 8:
            return f"({ddd}) {numero[:4]}-{numero[4:]}"
        return f"({ddd}) {numero}"
    return raw.strip()


def inteiro_seguro(v) -> str:
    """Converte float como 9999.0 → '9999', string '9999' → '9999'."""
    t = texto(v)
    if not t:
        return ""
    try:
        return str(int(float(t)))
    except Exception:
        return t


# ==============================
# LEITURA DA PLANILHA DO CLIENTE
# Colunas por POSIÇÃO:
#   A (0) → C / F
#   B (1) → CNPJ
#   C (2) → Descrição Conta (informativo — IGNORADO)
#   D (3) → Conta Patrimonial
# ==============================
def ler_planilha_cliente(arquivo_bytes: bytes) -> tuple:
    erros     = []
    registros = []

    try:
        try:
            df = pd.read_excel(
                io.BytesIO(arquivo_bytes),
                sheet_name=0, header=None, dtype=object,
            )
        except Exception:
            df = pd.read_excel(
                io.BytesIO(arquivo_bytes),
                sheet_name=0, header=None, dtype=object, engine="xlrd",
            )

        df = df.fillna("")

        # ── detecta linha de cabeçalho ──────────────────────
        inicio_dados = 1
        for i in range(min(10, len(df))):
            vals   = [str(v).strip().lower() for v in df.iloc[i].tolist()]
            joined = " ".join(vals)
            if any(k in joined for k in ["cnpj", "cpf", "c / f", "c/f", "tipo", "cliente", "fornecedor"]):
                inicio_dados = i + 1
                break

        for i in range(inicio_dados, len(df)):
            row = df.iloc[i].tolist()
            while len(row) < 4:
                row.append("")

            col_a = texto(row[0])   # C / F
            col_b = texto(row[1])   # CNPJ
            # col_c = ignorado       # Descrição Conta (informativo)
            col_d = texto(row[3])   # Conta Patrimonial

            if not col_a and not col_b and not col_d:
                continue

            cnpj_raw = so_numeros(col_b)
            if not cnpj_raw:
                continue

            tipo = col_a.strip().upper()
            if tipo not in ("C", "F"):
                tipo = "C"

            registros.append({
                "tipo"            : tipo,
                "cnpj_raw"        : cnpj_raw,
                "conta_patrimonial": col_d,
            })

    except Exception as e:
        erros.append(f"Erro ao ler planilha: {e}")

    return registros, erros


# ==============================
# CONSULTA API — RECEITA FEDERAL
# BrasilAPI (principal) + ReceitaWS (fallback)
# ==============================
BRASILAPI_URL = "https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
RECEITAWS_URL = "https://receitaws.com.br/v1/cnpj/{cnpj}"


def consultar_cnpj_api(cnpj_raw: str) -> dict:
    cnpj = so_numeros(cnpj_raw)
    if len(cnpj) != 14:
        return {"erro": "CNPJ deve ter 14 dígitos."}
    if not validar_cnpj(cnpj):
        return {"erro": "CNPJ inválido (dígitos verificadores incorretos)."}

    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}

    try:
        resp = requests.get(BRASILAPI_URL.format(cnpj=cnpj), headers=headers, timeout=15)
        if resp.status_code == 429:
            time.sleep(3)
            resp = requests.get(BRASILAPI_URL.format(cnpj=cnpj), headers=headers, timeout=15)
        if resp.status_code == 200:
            return _norm_brasilapi(resp.json())
    except Exception:
        pass

    try:
        resp = requests.get(RECEITAWS_URL.format(cnpj=cnpj), headers=headers, timeout=15)
        if resp.status_code == 200:
            return _norm_receitaws(resp.json())
    except Exception:
        pass

    return {"erro": "API indisponível. Tente novamente mais tarde."}


def _norm_brasilapi(data: dict) -> dict:
    if "message" in data or data.get("type") == "service_error":
        return {"erro": data.get("message", "CNPJ não encontrado.")}
    return {
        "cnpj"         : so_numeros(str(data.get("cnpj", ""))),
        "razao_social" : (data.get("razao_social") or "").strip(),
        "nome_fantasia": (data.get("nome_fantasia") or "").strip(),
        "situacao"     : (data.get("descricao_situacao_cadastral") or "").strip(),
        "uf"           : (data.get("uf") or "").strip(),
        "municipio"    : (data.get("municipio") or "").strip(),
        "bairro"       : (data.get("bairro") or "").strip(),
        "logradouro"   : (data.get("logradouro") or "").strip(),
        "numero"       : (data.get("numero") or "").strip(),
        "complemento"  : (data.get("complemento") or "").strip(),
        "cep"          : so_numeros(str(data.get("cep", ""))),
        "telefone"     : extrair_telefone(str(data.get("ddd_telefone_1", ""))),
        "fonte"        : "BrasilAPI",
    }


def _norm_receitaws(data: dict) -> dict:
    if data.get("status") == "ERROR":
        return {"erro": data.get("message", "CNPJ não encontrado.")}
    return {
        "cnpj"         : so_numeros(str(data.get("cnpj", ""))),
        "razao_social" : (data.get("nome") or "").strip(),
        "nome_fantasia": (data.get("fantasia") or "").strip(),
        "situacao"     : (data.get("situacao") or "").strip(),
        "uf"           : (data.get("uf") or "").strip(),
        "municipio"    : (data.get("municipio") or "").strip(),
        "bairro"       : (data.get("bairro") or "").strip(),
        "logradouro"   : (data.get("logradouro") or "").strip(),
        "numero"       : (data.get("numero") or "").strip(),
        "complemento"  : (data.get("complemento") or "").strip(),
        "cep"          : so_numeros(str(data.get("cep", ""))),
        "telefone"     : extrair_telefone(str(data.get("telefone", ""))),
        "fonte"        : "ReceitaWS",
    }


# ==============================
# MONTAGEM DO REGISTRO FINAL
# Regras exatas conforme especificação:
#   Col A → tipo       (planilha)
#   Col B → cnpj       (planilha)
#   Col C → IGNORADO
#   Col D → conta_patrimonial (planilha)
#   Demais campos → Receita Federal (API)
#   Insc. Estadual / Municipal / Conta Débito → em branco
# ==============================
def montar_registro(entrada: dict, api: dict) -> dict:
    tem_erro = "erro" in api
    return {
        # ── planilha do cliente ────────────────────────────────
        "cnpj"             : so_numeros(entrada["cnpj_raw"]),
        "_cnpj_fmt"        : formatar_cnpj(entrada["cnpj_raw"]),
        "tipo"             : entrada["tipo"],
        "conta_patrimonial": entrada.get("conta_patrimonial", ""),
        # ── em branco (não existem no modelo de entrada nem na API pública)
        "conta_debito"        : "",
        "inscricao_estadual"  : "",
        "inscricao_municipal" : "",
        # ── Receita Federal ────────────────────────────────────
        "razao_social" : api.get("razao_social", "") if not tem_erro else "",
        "nome_fantasia": api.get("nome_fantasia", "") if not tem_erro else "",
        "situacao"     : api.get("situacao", "")     if not tem_erro else "",
        "uf"           : api.get("uf", "")           if not tem_erro else "",
        "municipio"    : api.get("municipio", "")    if not tem_erro else "",
        "bairro"       : api.get("bairro", "")       if not tem_erro else "",
        "logradouro"   : api.get("logradouro", "")   if not tem_erro else "",
        "numero"       : api.get("numero", "")       if not tem_erro else "",
        "complemento"  : api.get("complemento", "")  if not tem_erro else "",
        "cep"          : api.get("cep", "")          if not tem_erro else "",
        "telefone"     : api.get("telefone", "")     if not tem_erro else "",
        # ── metadados internos ─────────────────────────────────
        "fonte"     : api.get("fonte", "—"),
        "_erro"     : tem_erro,
        "_msg_erro" : api.get("erro", ""),
    }


# ==============================
# GERAÇÃO DO EXCEL DE SAÍDA
# Espelho exato de planilhamodelodominio.xls
# 15 colunas + coluna "GERAR O ARQUIVO TXT" (em branco)
# ==============================
COLUNAS_SAIDA = [
    "C: cliente / F: Fornecedor",
    "Razão Social",
    "CPF/CNPJ",
    "Inscrição Estadual",
    "Inscrição Municipal",
    "UF",
    "Município",
    "Bairro",
    "Endereço",
    "Número Endereço",
    "Complemento Endereço",
    "CEP",
    "Telefone",
    "Conta débito",
    "Conta Patrimonial",
]


def gerar_excel_saida(registros: list) -> bytes:
    rows = []
    for r in registros:
        cnpj_num = so_numeros(r.get("cnpj", ""))
        cep_num  = so_numeros(r.get("cep", ""))

        # CNPJ e CEP como número inteiro (igual ao modelo original: 99999999000191.0)
        cnpj_val = float(cnpj_num) if cnpj_num else ""
        cep_val  = float(cep_num)  if cep_num  else ""

        rows.append({
            "C: cliente / F: Fornecedor": r.get("tipo", "C"),
            "Razão Social"              : r.get("razao_social", ""),
            "CPF/CNPJ"                  : cnpj_val,
            "Inscrição Estadual"        : inteiro_seguro(r.get("inscricao_estadual", "")),
            "Inscrição Municipal"       : inteiro_seguro(r.get("inscricao_municipal", "")),
            "UF"                        : r.get("uf", ""),
            "Município"                 : r.get("municipio", ""),
            "Bairro"                    : r.get("bairro", ""),
            "Endereço"                  : r.get("logradouro", ""),
            "Número Endereço"           : r.get("numero", ""),
            "Complemento Endereço"      : r.get("complemento", ""),
            "CEP"                       : cep_val,
            "Telefone"                  : r.get("telefone", ""),
            "Conta débito"              : inteiro_seguro(r.get("conta_debito", "")),
            "Conta Patrimonial"         : inteiro_seguro(r.get("conta_patrimonial", "")),
        })

    df = pd.DataFrame(rows, columns=COLUNAS_SAIDA)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Plan1")
        ws = writer.sheets["Plan1"]

        # larguras idênticas ao modelo original
        for col_letter, width in zip(
            list("ABCDEFGHIJKLMNO"),
            [28, 42, 20, 18, 18, 6, 26, 26, 36, 14, 22, 12, 20, 14, 16],
        ):
            ws.column_dimensions[col_letter].width = width

        header_fill  = PatternFill("solid", fgColor="FF8000")
        header_font  = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        thin_side    = Side(style="thin", color="CCCCCC")
        thin_border  = Border(
            left=thin_side, right=thin_side,
            top=thin_side, bottom=thin_side,
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align   = Alignment(horizontal="left",   vertical="center")
        row_fills    = [
            PatternFill("solid", fgColor="FFFFFF"),
            PatternFill("solid", fgColor="F5F5F5"),
        ]
        data_font = Font(name="Segoe UI", size=10)

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border

        for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
            for cell in row:
                cell.fill      = row_fills[ri % 2]
                cell.font      = data_font
                cell.alignment = left_align
                cell.border    = thin_border

        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf.read()


# ==============================
# GERAÇÃO DO TXT (cli_for.txt)
# Formato espelho de planilhamodelodominio.xls:
# campo1;campo2;...;campo15
# ==============================
def gerar_txt_saida(registros: list) -> bytes:
    """
    Gera o cli_for.txt com os 15 campos separados por ponto-e-vírgula,
    exatamente na mesma ordem das colunas do planilhamodelodominio.xls.
    """
    linhas = []
    for r in registros:
        cnpj_num = so_numeros(r.get("cnpj", ""))
        cep_num  = so_numeros(r.get("cep", ""))

        campos = [
            r.get("tipo", "C"),                               # C/F
            r.get("razao_social", ""),                        # Razão Social
            cnpj_num,                                         # CPF/CNPJ (só números)
            inteiro_seguro(r.get("inscricao_estadual", "")),  # Insc. Estadual
            inteiro_seguro(r.get("inscricao_municipal", "")), # Insc. Municipal
            r.get("uf", ""),                                  # UF
            r.get("municipio", ""),                           # Município
            r.get("bairro", ""),                              # Bairro
            r.get("logradouro", ""),                          # Endereço
            r.get("numero", ""),                              # Número
            r.get("complemento", ""),                         # Complemento
            cep_num,                                          # CEP (só números)
            r.get("telefone", ""),                            # Telefone
            inteiro_seguro(r.get("conta_debito", "")),        # Conta débito
            inteiro_seguro(r.get("conta_patrimonial", "")),   # Conta Patrimonial
        ]
        linhas.append(";".join(str(c) for c in campos))

    return ("\n".join(linhas) + "\n").encode("utf-8", errors="replace")


# ==============================
# MODELO DE ENTRADA PARA DOWNLOAD
# 4 colunas que o cliente preenche
# ==============================
def gerar_modelo_entrada() -> bytes:
    colunas = [
        "C: cliente / F: Fornecedor",
        "CPF/CNPJ",
        "Descrição Conta Plano Cliente",
        "Conta Patrimonial",
    ]
    exemplos = [
        ["C", "99.999.999/0001-91", "Clientes Nacionais",  "7777"],
        ["F", "88.888.888/0001-81", "Fornecedores Gerais", "6666"],
        ["C", "11.222.333/0001-81", "Clientes Exportação", "5555"],
    ]
    df  = pd.DataFrame(exemplos, columns=colunas)
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Clientes_Fornecedores")
        ws = writer.sheets["Clientes_Fornecedores"]

        for col_letter, width in zip("ABCD", [30, 22, 32, 18]):
            ws.column_dimensions[col_letter].width = width

        header_fill  = PatternFill("solid", fgColor="FF8000")
        header_font  = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        thin_side    = Side(style="thin", color="CCCCCC")
        thin_border  = Border(
            left=thin_side, right=thin_side,
            top=thin_side, bottom=thin_side,
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align   = Alignment(horizontal="left",   vertical="center")
        ex_fill      = PatternFill("solid", fgColor="FFF3E0")
        ex_font      = Font(name="Segoe UI", size=10, italic=True, color="888888")

        # cabeçalho
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border

        # col C com fundo diferente (informativo)
        ws["C1"].fill = PatternFill("solid", fgColor="777777")

        # linhas de exemplo
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill      = ex_fill
                cell.font      = ex_font
                cell.alignment = left_align
                cell.border    = thin_border

        ws.freeze_panes = "A2"

    buf.seek(0)
    return buf.read()


# ==============================
# PROCESSAMENTO CENTRAL
# ==============================
def processar(arquivo_bytes: bytes, log: list) -> list:
    registros_entrada, erros = ler_planilha_cliente(arquivo_bytes)

    for e in erros:
        log.append(f"⚠ {e}")

    if not registros_entrada:
        log.append("ERRO: Nenhum CNPJ válido encontrado na planilha.")
        return []

    log.append(f"📄 Planilha lida: {len(registros_entrada)} CNPJ(s) encontrado(s).")

    total  = len(registros_entrada)
    pbar   = st.progress(0, text="Consultando CNPJs na Receita Federal...")
    result = []

    for idx, entrada in enumerate(registros_entrada):
        cnpj_fmt  = formatar_cnpj(entrada["cnpj_raw"])
        tipo_desc = "Cliente" if entrada["tipo"] == "C" else "Fornecedor"

        pbar.progress(
            (idx + 1) / total,
            text=f"Consultando {cnpj_fmt} ({idx+1}/{total})...",
        )

        dados_api = consultar_cnpj_api(entrada["cnpj_raw"])
        registro  = montar_registro(entrada, dados_api)
        result.append(registro)

        if registro["_erro"]:
            log.append(f"✗ {cnpj_fmt} [{tipo_desc}] → {registro['_msg_erro']}")
        else:
            log.append(
                f"✓ {cnpj_fmt} [{tipo_desc}] → "
                f"{registro['razao_social']} [{registro['fonte']}]"
            )

        if idx < total - 1:
            time.sleep(0.8)

    pbar.empty()

    ok  = sum(1 for r in result if not r["_erro"])
    err = sum(1 for r in result if r["_erro"])
    log.append(f"✅ Sucesso: {ok}  |  ⚠ Erros: {err}  |  Total: {total}")

    return result


# ==============================
# INTERFACE STREAMLIT
# ==============================
def main():
    st.set_page_config(
        page_title="Domínio Sistemas | Cadastro Clientes/Fornecedores",
        page_icon="🟠",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    apply_tr_theme()

    # ── cabeçalho ──────────────────────────────────────────────
    st.markdown(
        f"""
        <div style="background:#444444; padding:24px 28px 18px 28px;
                    border-radius:8px; border-top:6px solid #FF8000;
                    margin-bottom:28px;">
            <h2 style="color:#FF8000; margin:0;
                       font-family:'Segoe UI',Arial,sans-serif;">
                🏢 Cadastro de Clientes / Fornecedores &nbsp;|&nbsp; {VERSAO}
            </h2>
            <p style="color:#DDDDDD; margin:6px 0 0 0;
                      font-family:'Segoe UI',Arial,sans-serif;">
                Importe a planilha com <strong>C/F</strong>,
                <strong>CNPJ</strong> e <strong>Conta Patrimonial</strong>,
                consulte a Receita Federal e exporte no formato
                <strong>Domínio Sistemas</strong>.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── sidebar ────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### 📥 Modelo de Planilha")
        st.markdown(
            "Baixe o modelo, preencha as **4 colunas** e faça o upload."
        )
        modelo_bytes = gerar_modelo_entrada()
        st.download_button(
            label="⬇ Baixar modelo de entrada (.xlsx)",
            data=modelo_bytes,
            file_name="modelo_clientes_fornecedores.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown("---")
        st.markdown("### 📋 Colunas do modelo")
        st.markdown(
            """
| Col | Campo | Uso |
|-----|-------|-----|
| **A** | C / F | Obrigatório |
| **B** | CNPJ | Obrigatório |
| **C** | Descrição Conta | ℹ Informativo |
| **D** | Conta Patrimonial | Obrigatório |
            """
        )
        st.markdown("---")
        st.markdown("### 📡 Fonte de dados")
        st.markdown(
            "Receita Federal via **BrasilAPI**, "
            "fallback automático para **ReceitaWS**."
        )
        st.markdown("---")
        st.markdown("### ℹ Sobre")
        st.markdown(f"**Versão:** {VERSAO}")
        st.markdown("**Thomson Reuters**")
        st.markdown("**Domínio Sistemas**")

    # ── instruções ────────────────────────────────────────────
    with st.expander("📖 **Instruções de Uso** — clique para expandir", expanded=False):
        st.markdown(
            """
            <div class="instrucoes-box">

            <h4>🔹 Passo 1 — Baixar o modelo</h4>
            <p>Clique em <b>⬇ Baixar modelo de entrada (.xlsx)</b>
            no menu lateral.</p>

            <h4>🔹 Passo 2 — Preencher a planilha</h4>
            <ul>
                <li><b>Coluna A</b> — <code>C: cliente / F: Fornecedor</code>:
                    digite <b>C</b> ou <b>F</b>.</li>
                <li><b>Coluna B</b> — <code>CPF/CNPJ</code>:
                    informe o CNPJ (com ou sem formatação).</li>
                <li><b>Coluna C</b> — <code>Descrição Conta Plano Cliente</code>:
                    campo <b>informativo</b>, não é processado nem exportado.</li>
                <li><b>Coluna D</b> — <code>Conta Patrimonial</code>:
                    informe o código da conta.</li>
            </ul>
            <p>Razão Social, Endereço, UF, Município, CEP, Telefone e demais
            dados são preenchidos <b>automaticamente</b> pela Receita Federal.</p>

            <h4>🔹 Passo 3 — Upload e processamento</h4>
            <ol>
                <li>Faça o upload da planilha preenchida.</li>
                <li>Clique em
                <b>▶ Consultar Receita Federal e Gerar Arquivos</b>.</li>
                <li>Aguarde (~1 segundo por CNPJ).</li>
            </ol>

            <h4>🔹 Passo 4 — Exportar</h4>
            <ul>
                <li><b>planilhamodelodominio.xlsx</b> →
                    15 colunas no formato exato do Domínio Sistemas.</li>
                <li><b>cli_for.txt</b> →
                    campos separados por <code>;</code>,
                    pronto para importação.</li>
            </ul>

            <h4>🔹 Passo 5 — Importar no Domínio</h4>
            <p>Contabilidade →
            <b>Utilitários → Importação → Clientes/Fornecedores</b>.</p>

            <hr>
            <h4>⚠ Observações</h4>
            <ul>
                <li>Inscrição Estadual/Municipal e Conta Débito ficam
                    <b>em branco</b> (não disponíveis via API pública).</li>
                <li>CNPJs com erro de consulta são exportados com os
                    campos da Receita Federal em branco.</li>
                <li>Limite da API: ~5 req/min —
                    intervalo automático aplicado.</li>
            </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("---")

    # ── estado da sessão ───────────────────────────────────────
    for key, default in [
        ("registros",   []),
        ("log",         [f"Aplicação pronta. Versão: {VERSAO}"]),
        ("xlsx_bytes",  None),
        ("txt_bytes",   None),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ── upload ─────────────────────────────────────────────────
    st.markdown("#### 📂 Upload da planilha preenchida")
    arquivo = st.file_uploader(
        "Selecione o arquivo (.xls / .xlsx)",
        type=["xls", "xlsx"],
        help="Planilha com 4 colunas: C/F | CNPJ | Descrição (ignorado) | Conta Patrimonial",
    )

    col1, col2 = st.columns([1, 1])
    with col1:
        processar_btn = st.button(
            "▶ Consultar Receita Federal e Gerar Arquivos",
            disabled=(arquivo is None),
            use_container_width=True,
            type="primary",
        )
    with col2:
        limpar = st.button("🗑 Limpar", use_container_width=True)

    if limpar:
        st.session_state.registros  = []
        st.session_state.log        = ["Campos limpos."]
        st.session_state.xlsx_bytes = None
        st.session_state.txt_bytes  = None
        st.rerun()

    # ── processamento ──────────────────────────────────────────
    if processar_btn and arquivo is not None:
        st.session_state.registros  = []
        st.session_state.log        = ["Iniciando processamento..."]
        st.session_state.xlsx_bytes = None
        st.session_state.txt_bytes  = None

        registros = processar(arquivo.read(), st.session_state.log)
        st.session_state.registros = registros

        if registros:
            st.session_state.xlsx_bytes = gerar_excel_saida(registros)
            st.session_state.txt_bytes  = gerar_txt_saida(registros)
            st.session_state.log.append("📁 Arquivos prontos para download.")

        st.rerun()

    # ── tabela de resultados ───────────────────────────────────
    if st.session_state.registros:
        st.markdown("---")
        st.markdown("#### 📊 Resultado da consulta")

        df_view = pd.DataFrame([
            {
                "CNPJ"             : r["_cnpj_fmt"],
                "Tipo"             : r["tipo"],
                "Razão Social"     : r["razao_social"],
                "Situação RF"      : r["situacao"],
                "UF"               : r["uf"],
                "Município"        : r["municipio"],
                "CEP"              : formatar_cep(r["cep"]) if r["cep"] else "",
                "Telefone"         : r["telefone"],
                "Conta Patrimonial": r["conta_patrimonial"],
                "Fonte"            : r["fonte"],
                "Status"           : "✅ OK" if not r["_erro"] else f"⚠ {r['_msg_erro']}",
            }
            for r in st.session_state.registros
        ])

        st.dataframe(
            df_view,
            use_container_width=True,
            hide_index=True,
            column_config={
                "CNPJ"         : st.column_config.TextColumn("CNPJ",         width="medium"),
                "Tipo"         : st.column_config.TextColumn("Tipo",         width="small"),
                "Razão Social" : st.column_config.TextColumn("Razão Social", width="large"),
                "Situação RF"  : st.column_config.TextColumn("Situação RF",  width="small"),
                "Status"       : st.column_config.TextColumn("Status",       width="large"),
            },
        )

        # ── métricas ──────────────────────────────────────────
        regs  = st.session_state.registros
        total = len(regs)
        ok    = sum(1 for r in regs if not r["_erro"])
        err   = total - ok
        cli   = sum(1 for r in regs if r["tipo"] == "C")
        forn  = sum(1 for r in regs if r["tipo"] == "F")

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Total",           total)
        c2.metric("✅ Sucesso",      ok)
        c3.metric("⚠ Erros",        err)
        c4.metric("👤 Clientes",     cli)
        c5.metric("🏭 Fornecedores", forn)

        if err:
            st.warning(
                f"⚠ {err} CNPJ(s) com erro de consulta — "
                "exportados com os campos da Receita Federal em branco."
            )

        # ── seção de downloads ────────────────────────────────
        st.markdown("---")
        st.markdown("#### ⬇ Exportar arquivos")

        col_a, col_b = st.columns(2)

        with col_a:
            st.markdown(
                """
                <div style="background:#F5F5F5; border:1px solid #E0E0E0;
                            border-left:4px solid #FF8000; border-radius:6px;
                            padding:14px 16px; margin-bottom:8px;">
                    <div style="font-weight:bold; color:#444; margin-bottom:4px;">
                        📊 planilhamodelodominio.xlsx
                    </div>
                    <div style="font-size:12px; color:#777;">
                        15 colunas no formato exato do Domínio Sistemas.<br>
                        Cabeçalho laranja · Zebrado · Linha congelada.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.session_state.xlsx_bytes:
                st.download_button(
                    label="⬇ Baixar planilhamodelodominio.xlsx",
                    data=st.session_state.xlsx_bytes,
                    file_name="planilhamodelodominio.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
            else:
                st.button(
                    "⬇ Baixar planilhamodelodominio.xlsx",
                    disabled=True,
                    use_container_width=True,
                )

        with col_b:
            st.markdown(
                """
                <div style="background:#F5F5F5; border:1px solid #E0E0E0;
                            border-left:4px solid #FF8000; border-radius:6px;
                            padding:14px 16px; margin-bottom:8px;">
                    <div style="font-weight:bold; color:#444; margin-bottom:4px;">
                        📄 cli_for.txt
                    </div>
                    <div style="font-size:12px; color:#777;">
                        Campos separados por <code>;</code> · 15 campos por linha.<br>
                        Pronto para importar no Domínio Sistemas.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            if st.session_state.txt_bytes:
                st.download_button(
                    label="⬇ Baixar cli_for.txt",
                    data=st.session_state.txt_bytes,
                    file_name="cli_for.txt",
                    mime="text/plain",
                    use_container_width=True,
                    type="primary",
                )
            else:
                st.button(
                    "⬇ Baixar cli_for.txt",
                    disabled=True,
                    use_container_width=True,
                )

    # ── log ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**Log de processamento**")
    log_texto = "\n".join(str(l) for l in st.session_state.log)
    tem_erro  = any(
        str(l).startswith("ERRO") or "✗" in str(l)
        for l in st.session_state.log
    )
    cor_borda = "#D32F2F" if tem_erro else "#388E3C"

    st.markdown(
        f"""
        <div style="background:#FCFCFC; border:1px solid {cor_borda};
                    border-radius:6px; padding:14px;
                    font-family:Consolas,monospace; font-size:13px;
                    white-space:pre-wrap; max-height:320px;
                    overflow-y:auto; color:#1F1F1F;">
{log_texto}
        </div>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
